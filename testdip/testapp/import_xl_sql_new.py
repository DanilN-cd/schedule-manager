import os
import re
from openpyxl import load_workbook
from testapp.models import Predmets, Group, PredM

# Регулярное выражение
pattern = r"(?<!ПМ)\.(\d{2}(?<!\.00)$)|(?<!ПМ)(\d{2}(?<!\.00)\*$)"  # .01 и .01* проходят, .00 не проходит

def check_symbols(s):
    pattern = r"(ДЗ|З)"  # Проверяем наличие "ДЗ" или "З"
    return 1 if re.search(pattern, s) else 0

def import_xl_sql(file_path, group_name, sheet_number):
    try:
        # Создаём или получаем группу из базы данных
        group, created = Group.objects.get_or_create(name=group_name)
        if created:
            print(f"Создана группа: {group.name}")
        else:
            print(f"Группа уже существует: {group.name}")
        print(f"Обработка файла: {file_path} для группы: {group.name}")

        # Загрузка файла Excel через openpyxl
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.worksheets[sheet_number]  # Выбираем указанный лист

        # Определяем колонки для чтения на основе первой цифры названия группы
        first_digit = int(group_name[0])
        column_mapping = {1: [8, 9], 2: [10, 11], 3: [12, 13], 4: [14, 15]}
        hour_columns = column_mapping.get(first_digit, [8, 9])  # По умолчанию [9, 10]

        # Обработка данных на листе
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
            cell_val = row[0]
            val = row[1]
            hour_1sem = row[hour_columns[0]]
            hour_2sem = row[hour_columns[1]]

            # Проверяем, что значения часов являются числами
            if isinstance(hour_1sem, (int, float)):
                hour_1sem = int(hour_1sem)
            else:
                hour_1sem = 0

            if isinstance(hour_2sem, (int, float)):
                hour_2sem = int(hour_2sem)
            else:
                hour_2sem = 0

            if hour_1sem == 0 and hour_2sem == 0:
                print(f"Пропуск предмета {val} (индекс: {cell_val}): оба значения часов равны нулю.")
                continue

            # Применяем регулярное выражение для проверки названия
            if isinstance(cell_val, str) and re.search(pattern, cell_val):
                predm, created = PredM.objects.get_or_create(name=val, ind=cell_val)

                if created:
                    print(f"Создан объект PredM: {predm.name}")
                else:
                    print(f"Объект PredM уже существует: {predm.name}")
                
                # Создаем или обновляем объект Predmets
                predmet, created = Predmets.objects.get_or_create(
                    name=predm,  # Указываем объект PredM как ForeignKey
                    group=group,  # Привязываем к текущей группе
                    defaults={
                        "hours_1sem": hour_1sem,
                        "hours_2sem": hour_2sem,
                        "hours_total": hour_1sem + hour_2sem,
                        "hours_remaining": hour_1sem + hour_2sem,
                        "pairs_remaining": (hour_1sem + hour_2sem) / 2,
                    }
                )

                if created:
                    print(f"Создан предмет: {predmet.name}")
                else:
                    # Если предмет уже существует, обновляем часы
                    predmet.hours_1sem = hour_1sem
                    predmet.hours_2sem = hour_2sem
                    predmet.hours_total = hour_1sem + hour_2sem
                    predmet.save()
                    print(f"Обновлен предмет: {predmet.name}")

        return True, "Файл успешно обработан"
    except Exception as e:
        print(f"Ошибка при обработке файла: {e}")
        return False, str(e)