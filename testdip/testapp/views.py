import json
import openpyxl
from openpyxl.styles import Alignment
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, Border, Side
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from testapp.models import Schedule, Group, Prepods, PredM, Cabs, Predmets
import datetime
from collections import defaultdict
from .forms import GroupForm, PrepodForm
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import FileSystemStorage
from .import_xl_sql_new import import_xl_sql
from .import_xl_sql_new_v2 import import_xl_sql_v2
from .algoritm import generate_schedule
from django.core.serializers.json import DjangoJSONEncoder

def run_algorithm_page(request):
    return render(request, 'run_algorithm.html')

def get_subject_id(request):
    subject_name_id = request.GET.get('subject_name_id')
    group_id = request.GET.get('group_id')

    predmet = Predmets.objects.filter(name_id=subject_name_id, group_id=group_id).first()

    if predmet:
        return JsonResponse({'subject_id': predmet.id})
    
    return JsonResponse({'error': 'Предмет не найден'}, status=400)

@csrf_exempt
def run_algorithm(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        start_date = datetime.datetime.strptime(data['start_date'], '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(data['end_date'], '%Y-%m-%d').date()
        generate_schedule(start_date, end_date)
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)

def upload_file(request):
    if request.method == 'POST' and request.FILES['file']:
        file = request.FILES['file']
        group_name = request.POST['group_name']
        sheet_number = int(request.POST['sheet_number']) - 1  # Преобразуем в индекс (начинается с 0)
        algorithm = request.POST['algorithm']  # Получаем выбранный алгоритм

        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        file_path = fs.path(filename)

        # Выбираем алгоритм на основе выбора пользователя
        if algorithm == 'old':
            success, message = import_xl_sql(file_path, group_name, sheet_number)
        elif algorithm == 'new':
            success, message = import_xl_sql_v2(file_path, group_name, sheet_number)
        else:
            return JsonResponse({'success': False, 'message': 'Неверный выбор алгоритма'}, status=400)

        if success:
            return JsonResponse({'success': True, 'message': message})
        else:
            return JsonResponse({'success': False, 'message': message})
    return JsonResponse({'success': False, 'message': 'Файл не загружен'}, status=400)

def index_page(request):
    num_list = [0, 1, 2, 3, 4, 5]
    num_pairs = [1, 2, 3, 4, 5, 6]
    all_dates = Schedule.objects.values_list('date', flat=True).distinct().order_by('date')

    if not all_dates:
        return render(request, "index.html", {"weekly_schedule": {}})

    weekly_schedule = defaultdict(lambda: defaultdict(list))

    for date in all_dates:
        week_start = date - datetime.timedelta(days=date.weekday())
        weekly_schedule[week_start][date] = list(Schedule.objects.filter(date=date).order_by('pair_number').select_related('subject', 'cabinet', 'group'))

    groups = Group.objects.all().prefetch_related('predmets_set').order_by('name')
    subjects = PredM.objects.all()
    cabinets = Cabs.objects.all().order_by('name')
    # Собираем информацию о преподавателях и их предметах для каждой строки (пары)
    
    predmets = list(Predmets.objects.values('id', 'name'))
    subjectss = list(PredM.objects.values('id', 'prepod'))

    return render(request, 'index.html', {
        'groups': groups,
        'weekly_schedule': dict(weekly_schedule),
        'num_list': num_list,
        'num_pairs': num_pairs,
        'subjects': subjects,
        'cabinets': cabinets,
        'predmets': json.dumps(predmets, cls=DjangoJSONEncoder),
        'subjectss': json.dumps(subjectss, cls=DjangoJSONEncoder)
    })
def get_group_subjects(request):
    group_id = request.GET.get('group_id')
    if not group_id:
        return JsonResponse({'success': False, 'error': 'group_id обязателен'}, status=400)

    predmets = Predmets.objects.filter(group_id=group_id).select_related('name')
    data = [
        {
            'id': predmet.id,
            'name': predmet.name.name
        }
        for predmet in predmets
    ]
    return JsonResponse({'success': True, 'subjects': data})

@csrf_exempt
def update_schedule(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            print("Полученные данные:", data)  # Логируем входные данные

            # Обработка обновлений
            for update in data.get('updates', []):
                entry = Schedule.objects.get(id=update['entryId'])
                if update['type'] == 'move':
                    if update['direction'] == 'up' and entry.pair_number > 1:
                        entry_above = Schedule.objects.get(date=entry.date, pair_number=entry.pair_number - 1, group=entry.group)
                        entry_above.pair_number += 1
                        entry_above.save()
                        entry.pair_number -= 1
                        entry.save()
                    elif update['direction'] == 'down' and entry.pair_number < 6:
                        entry_below = Schedule.objects.get(date=entry.date, pair_number=entry.pair_number + 1, group=entry.group)
                        entry_below.pair_number -= 1
                        entry_below.save()
                        entry.pair_number += 1
                        entry.save()
                else:
                    if update['type'] == 'subject':
                        entry.subject_id = update['value']
                    elif update['type'] == 'cabinet':
                        entry.cabinet_id = update['value']
                    entry.save()

            # Обработка удалений
            for entry_id in data.get('deletions', []):
                Schedule.objects.filter(id=entry_id).delete()

            return JsonResponse({'success': True})

        except Exception as e:
            print("Ошибка при обновлении расписания:", str(e))  # Логируем ошибку
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
        
@csrf_exempt
def delete_schedule_entry(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        entry_id = data.get('entryId')
        if entry_id:
            entry = Schedule.objects.get(id=entry_id)
            predmet = entry.subject

            # Обновляем поля Predmets
            predmet.hours_used -= 2
            predmet.hours_remaining += 2
            predmet.pairs_remaining += 1
            predmet.save()

            entry.delete()
            return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)

@csrf_exempt
def add_schedule_entry(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            print("Полученные данные:", data)  # Логируем входные данные

            # Проверяем наличие обязательных полей
            required_fields = ['date', 'pair_number', 'group_id', 'subject_id', 'cabinet_id']
            for field in required_fields:
                if field not in data:
                    return JsonResponse({'success': False, 'error': f'Отсутствует поле {field}'}, status=400)

            # Преобразуем строковые значения в числа
            pair_number = int(data['pair_number'])
            group_id = int(data['group_id'])
            subject_id = int(data['subject_id'])
            cabinet_id = int(data['cabinet_id'])

            # Преобразуем дату из строки в объект date
            date = datetime.datetime.strptime(data['date'], '%Y-%m-%d').date()

            # Определяем день недели
            weekday = date.strftime('%a')
            days_mapping = {'Mon': 'Пн', 'Tue': 'Вт', 'Wed': 'Ср', 'Thu': 'Чт', 'Fri': 'Пт', 'Sat': 'Сб'}
            weekday = days_mapping.get(weekday, '')

            # Проверяем, существуют ли объекты в БД
            group = Group.objects.get(id=group_id)
            subject = Predmets.objects.get(id=subject_id)
            cabinet = Cabs.objects.get(id=cabinet_id)

            # Обновляем поля Predmets
            subject.hours_used += 2
            subject.hours_remaining -= 2
            subject.pairs_remaining -= 1
            subject.save()

            # Создаём запись
            schedule_entry = Schedule.objects.create(
                date=date,
                weekday=weekday,
                pair_number=pair_number,
                group=group,
                subject=subject,
                cabinet=cabinet
            )

            print("Создана запись с ID:", schedule_entry.id)  # Логируем успешное создание

            return JsonResponse({'success': True, 'entry_id': schedule_entry.id})

        except Group.DoesNotExist:
            print("Ошибка: Группа не найдена")
            return JsonResponse({'success': False, 'error': 'Группа не найдена'}, status=400)
        except Predmets.DoesNotExist:
            print("Ошибка: Предмет не найден")
            return JsonResponse({'success': False, 'error': 'Предмет не найден'}, status=400)
        except Cabs.DoesNotExist:
            print("Ошибка: Кабинет не найден")
            return JsonResponse({'success': False, 'error': 'Кабинет не найден'}, status=400)
        except ValueError as e:
            print("Ошибка преобразования данных:", str(e))
            return JsonResponse({'success': False, 'error': f'Ошибка преобразования данных: {str(e)}'}, status=400)
        except Exception as e:
            print("Общая ошибка:", str(e))
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

@csrf_exempt
def move_schedule_entry(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        entry_id = data.get('entryId')
        direction = data.get('direction')
        if entry_id and direction:
            entry = Schedule.objects.get(id=entry_id)
            if direction == 'up' and entry.pair_number > 1:
                entry_above = Schedule.objects.get(date=entry.date, pair_number=entry.pair_number - 1, group=entry.group)
                entry_above.pair_number += 1
                entry_above.save()
                entry.pair_number -= 1
                entry.save()
            elif direction == 'down' and entry.pair_number < 6:
                entry_below = Schedule.objects.get(date=entry.date, pair_number=entry.pair_number + 1, group=entry.group)
                entry_below.pair_number -= 1
                entry_below.save()
                entry.pair_number += 1
                entry.save()
            return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)

def groups_page(request):
    groups = Group.objects.all()
    form = GroupForm()

    return render(request, 'groups.html', {'groups': groups, 'form': form})

def add_group(request):
    if request.method == 'POST':
        form = GroupForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('groups_page')
    else:
        form = GroupForm()
    return render(request, 'add_group.html', {'form': form})

def delete_group(request, group_id):
    group = get_object_or_404(Group, pk=group_id)
    group.delete()
    return redirect('groups_page')

def prep_page(request):
    prepods = Prepods.objects.all()
    form = PrepodForm()
    return render(request, 'prepods.html', {'prepods': prepods, 'form': form})

def add_prep(request):
    if request.method == 'POST':
        form = PrepodForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('prepods_page')
    else:
        form = PrepodForm()
    return render(request, 'add_prep.html', {'form': form})

def edit_prep(request, prepod_id):
    prepod = get_object_or_404(Prepods, pk=prepod_id)
    if request.method == 'POST':
        form = PrepodForm(request.POST, instance=prepod)
        if form.is_valid():
            form.save()
            return redirect('prepods_page')
    else:
        form = PrepodForm(instance=prepod)
    return render(request, 'edit_prep.html', {'form': form, 'prepod_id': prepod_id, 'prepod': prepod})

def delete_prep(request, prepod_id):
    prepod = get_object_or_404(Prepods, pk=prepod_id)
    prepod.delete()
    return redirect('prepods_page')

def home_page(request):
    return render(request, 'home.html')

def edit_group(request, group_id):
    group = get_object_or_404(Group, pk=group_id)
    if request.method == 'POST':
        form = GroupForm(request.POST, instance=group)
        if form.is_valid():
            form.save()
            return redirect('groups_page')
    else:
        form = GroupForm(instance=group)
    return render(request, 'edit_group.html', {'form': form, 'group_id': group_id, 'group': group})

def export_excel(request):
    days_ru = {
        "Monday": "понедельник",
        "Tuesday": "вторник",
        "Wednesday": "среда",
        "Thursday": "четверг",
        "Friday": "пятница",
        "Saturday": "суббота",
        "Sunday": "воскресенье"
    }
    # Получаем дату начала недели из запроса
    week_start_str = request.GET.get("week_start")
    week_start = datetime.datetime.strptime(week_start_str, "%Y-%m-%d").date()

    # Получаем расписание на выбранную неделю
    weekly_schedule = defaultdict(lambda: defaultdict(list))
    all_dates = [week_start + datetime.timedelta(days=i) for i in range(6)]
    for date in all_dates:
        weekly_schedule[week_start][date] = list(Schedule.objects.filter(date=date).order_by('pair_number'))

    groups = Group.objects.all().order_by('name')

    # Создаем новый Excel-файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Расписание"

    # Форматирование границ
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Устанавливаем ширину столбцов
    ws.column_dimensions['A'].width = 7  # Дата
    ws.column_dimensions['B'].width = 3  # День недели
    ws.column_dimensions['C'].width = 5  # № пары

    for i in range(len(groups)):
        ws.column_dimensions[chr(68 + i * 2)].width = 15  # D, F, H... - Предмет
        ws.column_dimensions[chr(69 + i * 2)].width = 4   # E, G, I... - Кабинет

    # Заголовки для даты и номера пары
    ws.merge_cells('A6:B6')
    ws['A6'] = 'Дата'
    ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A6'].font = Font(bold=True)

    ws['C6'] = '№ пары'
    ws['C6'].alignment = Alignment(horizontal='left', vertical='center')
    ws['C6'].font = Font(bold=True)

    # Заголовки для групп
    col = 4  # D - первая колонка для первой группы
    for group in groups:
        ws.cell(row=6, column=col, value='Группа №')
        ws.cell(row=6, column=col).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=6, column=col).font = Font(bold=True)

        ws.cell(row=6, column=col+1, value=group.name)
        ws.cell(row=6, column=col+1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=6, column=col+1).font = Font(bold=True)

        col += 2

    # Заполнение данных (начиная с 6 строки)
    row = 7
    for date, schedule_for_date in weekly_schedule[week_start].items():
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 5, end_column=1)
        ws.merge_cells(start_row=row, start_column=2, end_row=row + 5, end_column=2)

        ws.cell(row=row, column=1, value=days_ru[date.strftime("%A")])
        ws.cell(row=row, column=2, value=date.strftime("%d.%m.%Y"))
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center', textRotation=90)

        # Заполнение пар
        for pair_number in range(1, 7):
            ws.cell(row=row, column=3, value=pair_number)
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')

            col = 4
            for group in groups:
                entry = next((e for e in schedule_for_date if e.pair_number == pair_number and e.group == group), None)
                if entry:
                    ws.cell(row=row, column=col, value=entry.subject.name.name)
                    ws.cell(row=row, column=col+1, value=entry.cabinet.name)
                col += 2

            row += 1

    # Отправка файла
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Расписание_{week_start_str}.xlsx"'
    wb.save(response)
    return response
